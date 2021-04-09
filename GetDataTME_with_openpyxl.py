# -*- coding: utf-8 -*-
"""
Created on Wed Nov 11 10:27:28 2020

Получение данных о электронных компонентах с сайта TME через его API
Наименования (Артикулы) помещаются в первый столбец файла XLSX, без пропусков строк, можно с производителем через пробел.
5760 шт в час примерно
"""
import openpyxl,time
from TME_Python_API import product_import_tme


class GetData:
    """Класс для получения информации по наименованию детали сайта TME"""
    # переменная-флаг для принудительного выхода из цикла поиска, используется графическим интерфейсом.
    flag=1
    
    @staticmethod
    # Функция считает все ячейки первого столбца в которых что то написано,до тех пор пока не встретит пустую ячейку "None"
    def number_of_articles(path):
        """Функция считает все ячейки первого столбца в которых что то написано, до тех пор пока не встретит пустую ячейку None"""
        # Открываем Эксель
        wb = openpyxl.load_workbook(path)#путь к файлу
        sheet = wb.active
        #выясняем количество артикулов в файле эксель
        i = 1
        while sheet['A'+str(i)].value != None:
            i+=1
        wb.save(path)
        return i-1
    @staticmethod
    # Функция создающая список артикулов. Принимает число артикулов и номер колонки в виде буквы (str): 'A'- первая колонка
    #
    def articlesList(n,column,path):
        """ Функция создающая список артикулов. Принимает число артикулов и номер колонки в виде буквы (str): 'A'- первая колонка"""
        # Открываем Эксель
        wb = openpyxl.load_workbook(path)#путь к файлу
        sheet = wb.active 
        cord_in = column+str(1)
        cord_out = column+str((n)) 
        # формирование списка артикулов
        vals = [v[0].value for v in sheet[cord_in : cord_out]] #[r[0].value for r in sheet.Range(cord)]
        return vals
        wb.save(path)
    
    def __init__(self, xlsxpath, parameters, token, app_secret):
        self.path = xlsxpath
        self.params = parameters
        self.token = token
        self.app_secret = app_secret

        self.action0 = 'Products/GetProducts' 
        self.action1 = 'Products/Search'
        self.action2 = 'Products/GetParameters'
        self.action3 = 'Products/GetProductsFiles'
        self.params_key1 = 'SymbolList[0]'
        self.params_key2 = 'SearchPlain'
        
        self.quantity_of_articles = GetData.number_of_articles(self.path)
        self.articles_list = GetData.articlesList(self.quantity_of_articles,"A",self.path)
        
    def get_data(self, part_name, action, params_key):
        """Получение массива от TME по экшену. Отдает None в случае неверного запроса"""
        prms = self.params.copy()
        prms[params_key] = part_name
        try:
            all_data = product_import_tme(self.token, self.app_secret, action, prms)
            return all_data
            
        except Exception:
            all_data = None
            print
            return all_data
   
    # Функция для вывода пинга
    #   
    def ping(self):
        """Функция для вывода пинга"""
        ping_data = self.get_data( None,'Utils/Ping',None)
        return ping_data
        
    def get_Search(self, part_name):
        """Поиск по экшену Search"""
        return self.get_data( part_name, self.action1, self.params_key2)
    
    def get_Products(self, part_name):
        """Поиск по экшену GetProducts"""
        return self.get_data( part_name, self.action0, self.params_key1)
    
    def get_Parameters(self, part_name):
        """Поиск по экшену GetParameters"""
        return self.get_data( part_name, self.action2, self.params_key1)
    
    def get_ProductsFiles(self, part_name):
        """Поиск по экшену GetProductsFiles"""
        return self.get_data( part_name, self.action3, self.params_key1)
    
    def get_dict(self,part_name):
        """Функция использеет экшены, формирует и отдает словарь для одной детали. Получает на входе артикул детали."""
        line_data = {"Symbol": "","Description": None,"Weight": None,"Photo": None,"ProductInformationPage": None,"OriginalSymbol": "","Producer":None, "ParameterList": "", "DocumentUrl": None}
        piece_of_data = self.get_Products(part_name)
        if piece_of_data == None:      
            try:
                print("Поиск детали",part_name,"через экшн Search")
                piece_of_data = self.get_Search(part_name)
                part_name = piece_of_data['Data']['ProductList'][0]['Symbol']
                
            except IndexError:
                print("Детали", part_name, "нет на TME")
                piece_of_data = None
            else:
                piece_of_data = self.get_Products(part_name)
        
        if piece_of_data != None and GetData.flag == 1:           
            print("Поиск по экшену GetProducts")
            try:    
                line_data["Symbol"] = piece_of_data['Data']['ProductList'][0]['Symbol']
            except TypeError: #IndexError:
                print(part_name, ": Название детали не найдено")
            try:
                line_data["Description"] = piece_of_data['Data']['ProductList'][0]['Description']
            except IndexError:
                print(part_name, ": Нет дескрипшина")
            try:
                photo = "https:" + piece_of_data['Data']['ProductList'][0]["Photo"]
                if photo != "https:":
                    line_data["Photo"] = "https:" + piece_of_data['Data']['ProductList'][0]["Photo"]
                else:
                    line_data["Photo"] = ""
            except IndexError:
                print(part_name, ": Нет фото")    
            try:    
                line_data["ProductInformationPage"] = piece_of_data['Data']['ProductList'][0]["ProductInformationPage"]
            except IndexError:
                print(part_name, ": Нет ссылки на деталь")    
            try:   
                line_data["OriginalSymbol"] = piece_of_data['Data']['ProductList'][0]["OriginalSymbol"]
            except IndexError:
                print(part_name, ": Нет оригинального названия")   
            try:   
                line_data["Producer"] = piece_of_data['Data']['ProductList'][0]["Producer"]
            except IndexError:
                print(part_name, ": Нет производителя")      
            try:   
                weight = piece_of_data['Data']['ProductList'][0]["Weight"]
                
            except IndexError:
                print(part_name,": Нет веса" )
            else:
                if piece_of_data['Data']['ProductList'][0]['WeightUnit']=='g':
                    weight = weight*0.001
                    #print('{:f}'.format(weight))
                    weight = round(weight,(('{:f}'.format(weight)).count('0'))+1) # округление
                    line_data["Weight"]= weight
                elif piece_of_data['Data']['ProductList'][0]['WeightUnit']=='mg':
                    #print("Вес в МГ")
                    weight = weight*0.000001
                    weight = round(weight,(('{:f}'.format(weight)).count('0'))+1) # округление
                    line_data["Weight"]= weight
                else:
                    line_data["Weight"]= weight
            if GetData.flag == 1:
                print("Получение параметров детали",part_name,". Экшн GetParameters")
                try:
                    piece_of_data = self.get_Parameters(part_name)
                except Exception:
                    print("Неправильное название детели или нет доступа к TME")
                else:
                    try:
                        print(piece_of_data['Data']['ProductList'][0]['ParameterList'][1]['ParameterName'], piece_of_data['Data']['ProductList'][0]['ParameterList'][1]['ParameterValue'])
                    except IndexError:
                        print("\nОшибка структуры ответа\n")
                    else:
                        parameters = {}
                        for i in piece_of_data['Data']['ProductList'][0]['ParameterList']:
                            parameters[i['ParameterName']] = i['ParameterValue']
                        parameters.pop('Производитель','Нет ключа производитель')
                        parameters.pop('#Promotion','Нет ключа')
                        parameters.pop('#Promotion','Нет ключа')
                        parameters.pop('вес','Нет ключа')
                        parameters.pop('Вес','Нет ключа')
                        parameters.pop('Сопутствующие товары','Нет ключа')
                        parameters.pop('Ресурс','Нет ключа')
                        parameters.pop('Дополнительные функции','Нет ключа')
                        parameters.pop('Электрический ресурс','Нет ключа')
                        parameters.pop('Alias','Нет ключа')
                        parameters.pop('Соответствуют норме','Нет ключа')
                        line_data["ParameterList"] = parameters
            if GetData.flag == 1:
                print("Получение даташита детали",part_name,". Экшн GetProductsFiles")
                try:
                    piece_of_data = self.get_ProductsFiles(part_name)
                except Exception:
                    print("Неправильное название детели или нет доступа к TME")
                else:
                    try:
                        line_data["DocumentUrl"] = "https:"+ piece_of_data['Data']['ProductList'][0]['Files']['DocumentList'][0]['DocumentUrl']
                    except IndexError:
                        print("Нет даташита")
                    if line_data["Photo"] == None:
                        try:
                            line_data["Photo"] = "https:"+ piece_of_data['Data']['ProductList'][0]['Files']["PhotoList"]
                        except IndexError:
                            line_data["Photo"] = None
                       
        return line_data
            
            
    def duty_cycle(self, rng1=0):
        """ОСНОВНОЙ ЦИКЛ. Цикл проходящий по всем деталям из файла XLSX, заполняющая все графы."""
        self.quantity_of_articles = GetData.number_of_articles(self.path)
        self.articles_list = GetData.articlesList(self.quantity_of_articles,"A",self.path)
        # Открываем Эксель
        wb = openpyxl.load_workbook(self.path)#путь к файлу
        sheet = wb.active
        rng2=len(self.articles_list)
        index = 0
        for j in range(rng1,rng2):
            #global GetData.flag
            if GetData.flag != 1:
                break
            else:
                index += 1
                print('#',index)
                if self.articles_list[j] != None:
                    line_data = self.get_dict(self.articles_list[j])
                    sheet['B'+str(j+1)] = str(line_data["Symbol"])
                    sheet['C'+str(j+1)] = line_data["Description"]
                    sheet['D'+str(j+1)] = line_data["Weight"]
                    sheet['E'+str(j+1)] = line_data["Photo"]
                    sheet['F'+str(j+1)] = line_data["ProductInformationPage"]
                    sheet['G'+str(j+1)] = str(line_data["ParameterList"])
                    sheet['H'+str(j+1)] = line_data["DocumentUrl"]
                    sheet['K'+str(j+1)] = line_data["Producer"]
                    sheet['L'+str(j+1)] = str(line_data["OriginalSymbol"])
        wb.save(self.path)
        print("Цикл обращений к TME завершен.")
                                          
# Испытания      
if __name__ == '__main__':
    xlsxpath = "productdata.xlsx"
    params={'Country' : 'RU','Language' : 'RU',}
    token = 'ac434c181917ed4e51c49a2027bfd040e9f2da0054be7'
    app_secret = '0b748f6e5d340d693703'
    
    A = GetData(xlsxpath,params,token,app_secret)
    A.duty_cycle()
    

    
    
    
    