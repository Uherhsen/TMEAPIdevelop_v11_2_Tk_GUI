# -*- coding: utf-8 -*-
"""
Created on Tue Apr 14 20:40:50 2020
"""
import description_assembly as da
#from GetDataTME_with_openpyxl import number_of_articles, articlesList, search_articles,search_param, products_files
from GetDataTME_with_openpyxl import GetData as GD 
from tkinter import *
from tkinter import messagebox as mb
import openpyxl, threading, json, os, shutil, webbrowser, configparser
#
# ПЕРЕМЕННЫЕ ДЛЯ ФУНКЦИЙ GetDataTME_with_openpyxl
#
xlsxpath = "productdata.xlsx"
jsonfilename = 'database.json'
params={'Country' : 'RU','Language' : 'RU',}
token = 'ac434c181917ed4e51c49a2027bfd040e9f2da0054be7'
app_secret = '0b748f6e5d340d693703'
action1 = 'Products/Search' # request method, метод пинг Utils/Ping
action2 = 'Products/GetParameters'
action3 = 'Products/GetProductsFiles'
GD.flag = 0
get_data = GD(xlsxpath,params,token,app_secret)
TEXT=''
# .ini
config = configparser.ConfigParser()
config.read("config.ini")
app_geometry = config['APP']['geometry']
clmn_key1 = config['COLUMNS']['арт']
clmn_key2 = config['COLUMNS']['описание']
clmn_key3 = config['COLUMNS']['параметры']
clmn_key4 = config['COLUMNS']['вес']
clmn_key5 = config['COLUMNS']['доки']
clmn_key6 = config['COLUMNS']['картинки']
clmn_key7 = config['COLUMNS']['краткопе']
clmn_key8= config['COLUMNS']['типы']
clmn_key9= config['COLUMNS']['ссылкаопя']
#
# ФУНКЦИИ = = = = = = = = = = = = = = = = = = = = = = = = = = = ==  = = = = = = = ===  !
#
# Выход
def quitme():
    cancelDATA()
    root.update_idletasks() # обновляем данные об окне
    config['APP']['geometry'] = root.geometry()
    if lbox.winfo_ismapped():
        config['COLUMNS']['арт'] = '1'
    else:
        config['COLUMNS']['арт'] = '0'
    clmn2.visibility_config()
    clmn3.visibility_config()
    clmn4.visibility_config()
    clmn5.visibility_config()
    clmn6.visibility_config()
    clmn7.visibility_config()
    clmn8.visibility_config()
    clmn9.visibility_config()
    with open('config.ini', 'w') as configfile:    # save
        config.write(configfile)
    root.destroy()
#
# Скрыть артикулы
def visibilityArt():
    if lbox.winfo_ismapped():
        Art_visibility_button['text']="Артикулы"
        Art_visibility_button['bg']='grey90'
        label_art.pack_forget()
        empty_label.pack_forget()
        lbox.pack_forget()
        scroll.pack_forget()
        scrollx.pack_forget()
        f1.pack_forget()
        f1_1.pack_forget()
    else:
        Art_visibility_button['text']="Арт.[x]"
        Art_visibility_button['bg']='grey95'
        f1.pack(side=LEFT)
        f1_1.pack(side=LEFT,ipady=8)
        label_art.pack(side=TOP)
        lbox.pack()
        scrollx.pack(fill=X, side=BOTTOM, ipadx=50, padx=2)
        scroll.pack(side=TOP, anchor=N, ipady=178)
        empty_label.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)
#
# Словарь из БД
def json_data(jsonfilename):
    try:
        data_list = json.load(open(jsonfilename)) # FileNotFoundError:
        return data_list
    except FileNotFoundError:
        print('Файл базы database.json отсутствует, или неверно указанно имя файла')
# Функция для замены элемента в словаре None на ""
def NoneRemove(mylist):
    if mylist.count(None) > 0:
        for i in range(mylist.count(None)):
            None_index=mylist.index(None)
            mylist.remove(None)
            mylist.insert(None_index,'')
    else:
        pass

# ОБновление колонок
def get_xlsxdata():
    # ПОЛУЧАЕМ КОЛИЧЕСТВО АРТИКУЛОВ И ИХ СПИСОК
    n = GD.number_of_articles(xlsxpath)
    #print("Кол-во артикуло",n)
    work_articles_list_A = GD.articlesList(n, 'A', xlsxpath)
    work_articles_list_I = GD.articlesList(n, 'I', xlsxpath)
    work_articles_list_G = GD.articlesList(n, 'G', xlsxpath)
    #work_articles_list_D = GD.articlesList(n, 'D', xlsxpath)
    # Убираем отображение весов с экспонентой 
    pre_work_articles_list_D = GD.articlesList(n, 'D', xlsxpath)
    work_articles_list_D=[]
    for i in pre_work_articles_list_D:
        try:
            i='{:f}'.format(i)
            i=i.rstrip('0')
        except TypeError:
            i=None
        work_articles_list_D.append(i)
    #work_articles_list_D = ['{:f}'.format(v) for v in work_articles_list_D] # TypeError
    work_articles_list_H = GD.articlesList(n, 'H', xlsxpath)
    work_articles_list_E = GD.articlesList(n, 'E', xlsxpath)
    work_articles_list_C = GD.articlesList(n, 'C', xlsxpath)
    work_articles_list_J = GD.articlesList(n, 'J', xlsxpath)
    work_articles_list_link = []
    data_json= json_data(jsonfilename)
    for i in work_articles_list_J:
        try:
            work_articles_list_link.append(data_json[0][i])
        except KeyError:
            work_articles_list_link.append(None)
            continue   
    #work_articles_list_link = [ data_json[0][i] for i in work_articles_list_J]
    # Отчистка ПОЛЕЙ ++++++=============+==+ == = = = =  = =========================>
    lbox.delete(0,END)
    lbox1_descr.delete(0,END)
    lbox2_param.delete(0,END)
    lbox3_weight.delete(0,END)
    lbox4_datasheet.delete(0,END)
    lbox5_picture.delete(0,END)
    lbox6_specification.delete(0,END)
    lbox7_type.delete(0,END)
    lbox8_link.delete(0,END)
    # Замена None на "", что бы создать пустые строки
    NoneRemove(work_articles_list_I)
    NoneRemove(work_articles_list_G)
    NoneRemove(work_articles_list_D)
    NoneRemove(work_articles_list_H)
    NoneRemove(work_articles_list_E)
    NoneRemove(work_articles_list_C)
    NoneRemove(work_articles_list_J)
    NoneRemove(work_articles_list_link)
    # ЗАПОЛНЕНИЕ ПОЛЕЙ ========= = = = = = == = = = = = = = ======== = = = = == = = >
    # ревер списков
    work_articles_list_A.reverse()
    work_articles_list_I.reverse()
    work_articles_list_G.reverse()
    work_articles_list_D.reverse()
    work_articles_list_H.reverse()
    work_articles_list_E.reverse()
    work_articles_list_C.reverse()
    work_articles_list_J.reverse()
    work_articles_list_link.reverse()
    # Вставляем значения
    for i in work_articles_list_A: lbox.insert(0,i)
    for i in work_articles_list_I: lbox1_descr.insert(0,i)
    for i in work_articles_list_G: lbox2_param.insert(0,i)
    for i in work_articles_list_D: lbox3_weight.insert(0,i)
    for i in work_articles_list_H: lbox4_datasheet.insert(0,i)
    for i in work_articles_list_E: lbox5_picture.insert(0,i)
    for i in work_articles_list_C: lbox6_specification.insert(0,i)
    for i in work_articles_list_J: lbox7_type.insert(0,i)
    for i in work_articles_list_link: lbox8_link.insert(0,i)
# ОБРАБОТКА АРТИКУЛОВ 
def startDATA(): #
    # Кусок кода для проверки через кол-во потоков 
    '''
    print('\n\nthreading.activeCount()=',threading.activeCount(),'\n\n')
    if threading.activeCount() < 7:
        data_thread = threading.Thread( target = my_thread, daemon=True)
        data_thread.start()
        print('\n\nthreading.activeCount()=',threading.activeCount(),'\n\n')
    else:
        print('\n\nУже в работе\n\nthreading.activeCount()=',threading.activeCount(),'\n\n') '''
    # Рабочий кусок кода через глобальный флаг
    if GD.flag == 0:
        data_thread = threading.Thread( target = my_thread, daemon=True)
        data_thread.start()
        
        start_button['text']= "Стоп"
        start_button['bg']='grey95'
        
    else:
        cancelDATA()
        start_button['text']= "Запуск"
        start_button['bg']='brown1'
        
        
# Поток
def my_thread():
    GD.flag = 1
    get_data.duty_cycle() # ____________________________________________________THEAD PROCESS______________________________________________________________
    
    # попытки выводить в лейлбл информацию
    #infolabel['text']=TEXT
    #infolabel.update_idletasks()
    
    # функции описания
    if GD.flag == 1:
        da.checkKey(xlsxpath,jsonfilename)
        da.desc_assembly(xlsxpath,jsonfilename)
    # Обновление столбцов
    get_xlsxdata()
    root.update_idletasks()
    start_button['text']= "Запуск"
    start_button['bg']='brown1'
    if GD.flag == 1:
        mb.showinfo("Завершено","Обработка завершена")
    GD.flag = 0 # сбрасываем флаг на 0
#      
# Запуск по нажатию Enter
def startDATAenter(event):
    startDATA()
#
# Отмена процесса
def cancelDATA():
    GD.flag = 0
    print('\n\nОТМЕНА\n\n')
# Сборка ОПИСАНИЙ 
def startASSEMBLY():
    da.checkKey(xlsxpath,jsonfilename)
    da.desc_assembly(xlsxpath,jsonfilename)
    get_xlsxdata()
    root.update_idletasks()
# РЕДАКТОР БАЗЫ ДАННЫХ
def DBedit():
    if GD.flag == 0:
        os.startfile(r'database.json')
    else:
        mb.showwarning("В работе!","Нельзя редактировать БД пока запущен рабочий цикл!")
# ОТКРЫТИЕ РАБОЧЕГО XLSX
def editXLSX():
    if GD.flag == 0:
        os.startfile(r'productdata.xlsx')
    else:
        mb.showwarning("В работе!","Нельзя редактировать XLSX пока запущен рабочий цикл!")
# КНОПКА УДАЛИТЬ ВСЕ
def delAll():
    #lbox.delete(0,END)
    if GD.flag == 0:
        # И УДАЛЯЕМ ИЗ ЭКСЕЛЯ, точнен просто создаем новый, пустой
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        wb.save( "productdata.xlsx")
        get_xlsxdata()
        root.update_idletasks()
    else:
        mb.showwarning("В работе!","Нельзя отчистить данные пока запущен рабочий цикл!")
# ВСТАВКА Ctrl+V
def ctrlv(event):
    art=root.clipboard_get()
    art=art.rstrip()
    art=art.split('\n')
    art.reverse()
    select = list(lbox.curselection())
    try:
        last_select_value=select[-1]
    except Exception:
        last_select_value=-1
    for i in art:
        lbox.insert((last_select_value+1),i)
    new_values=lbox.get(0,END)
    wb2 = openpyxl.Workbook()
    #ws2 = wb2.create_sheet()
    ws2=wb2.active
    #ival=-1
    for row in range(1,(len(new_values)+1)):
        #ival+=1
        ws2.cell(column=1, row=row, value=new_values[row-1])
    #ws2['B1']='ТЕСТ!!!'
    wb2.save(xlsxpath)
    get_xlsxdata()
    root.update_idletasks()
# Удаления как из поля программы так и из файла выделенных артикулов
def delList(event):
    select = list(lbox.curselection())
    select.reverse()
    for i in select:
        #lbox.delete(i)
        wb = openpyxl.load_workbook(xlsxpath)
        ws = wb.active
        ws.delete_rows(i+1)
        wb.save(xlsxpath)
    get_xlsxdata()
    root.update_idletasks()
# Комбинация клавишь "Вырезать" CTRL+X
def ctrlx(event):
    root.clipboard_clear()
    select = list(lbox.curselection())
    if select == []:
        lbox.select_set(0,END)
        select = list(lbox.curselection())
    clipget=list(map(lbox.get,select))
    #newclipget=list(map(lambda x: str(x) +' \n', clipget)) # ДОБАВЛЯЕМ ПЕРЕНОС ДЛЯ РАЗДЕЛЕНИЯ ДАННЫХ (слишком мудренно)
    newclipget=str(clipget)[1:-1]
    newclipget=newclipget.replace("'","")
    newclipget=newclipget.replace(", ","\n")
    root.clipboard_append(newclipget)
    select.reverse()
    for i in select:
        #lbox.delete(i)
        wb = openpyxl.load_workbook(xlsxpath)
        ws = wb.active
        ws.delete_rows(i+1)
        wb.save(xlsxpath)
    get_xlsxdata()
    root.update_idletasks()
# Функция привязки вертикального скрола кнескольким виджетам
def on_scrollbar(*args):
    lbox.yview(*args)
    lbox1_descr.yview(*args)
    lbox2_param.yview(*args)
    lbox3_weight.yview(*args)
    lbox4_datasheet.yview(*args)
    lbox5_picture.yview(*args)
    lbox6_specification.yview(*args)
    lbox7_type.yview(*args)
    lbox8_link.yview(*args)
#======================== Раскрытие меню ссылок на описание
def links_menu(event):
    pos=lbox7_type.curselection()    
    if pos == ():
        mb.showinfo('Не выбран элемент','Для открытия меню нужно выделить элемент левой кнопкой мыши')
        print('Не выбран элемент')
    else:
        linksmenu.post(event.x_root,event.y_root)
        #print(lbox8_link.get(pos))
# ================================================ Вызов меню ссылок        
def links_menu_command():
    # выход из топлевела
    def exitmenu(event):
        toplist.destroy()
    # создаем окно    
    toplist=Toplevel()
    x = root.winfo_pointerx()
    y = root.winfo_pointery()
    #print(x,y)
    toplist.geometry('180x300'+'+'+str(x)+'+'+str(y))
    # Убирает рамку вокруг окна
    toplist.overrideredirect(True) 
    linksListbox=Listbox(toplist, width=27, height=12,bg='grey96')
    # linksListbox.delete(0,END)
    # получаем сразу позицию из колонки с типами
    pos=lbox7_type.curselection()
    TMEtype=lbox7_type.get(pos)
    # Закидываем из базы список ссылок на описания
    try:
        data = json.load(open(jsonfilename)) # FileNotFoundError:
    except FileNotFoundError:
        linksListbox.insert(0,'Файл базы database.json отсутствует, или неверно указанно имя файла')
        print('Файл базы database.json отсутствует, или неверно указанно имя файла')
    links=[*data[1]]
    links.sort() # ------------------------------------------------------------------------------------------ сортировка списка ссылок на описание
    links.reverse()
    for i in links: linksListbox.insert(0,i)
    linksListbox.insert(0,' - ')
    # Замена ссылки для типа        
    def newLink(event):
        #print(data[0][TMEtype])
        indexlinksListbox = linksListbox.curselection() # номер в сплывающем списке
        if indexlinksListbox == ():
            mb.showinfo('Не выбран элемент','Для открытия меню нужно выделить элемент левой кнопкой мыши')
            print('Не выбрана ссылка')
        else:
            newlink = linksListbox.get(indexlinksListbox)
            #print('\n\n',TMEtype,' - ', newlink,'\n\n')
            mes = 'Установить для "'+ str(TMEtype) +'" ссылку на описание: "'+ str(newlink)+'" ?'
            answer = mb.askyesno(title='Изменение ссылки на описание', message=mes)
            if answer == True:
                if newlink == ' - ':
                    data[0][TMEtype] = None
                else:
                    data[0][TMEtype] = newlink
                #print(data[0][TMEtype])
                
                # Создаем копию БД JSON
                jsonfilename_backup_copy = jsonfilename + '_backup_copy.json'
                if os.path.exists(jsonfilename_backup_copy):
                    os.remove(jsonfilename_backup_copy)
                    shutil.copy2(jsonfilename, jsonfilename_backup_copy)
                else:
                    shutil.copy2(jsonfilename, jsonfilename_backup_copy)
                
                # Загружаем обновленный словарь в базу json
                with open(jsonfilename,'w') as file:
                    json.dump(data, file, indent=2, ensure_ascii=False)
                # Сборка ОПИСАНИЙ
                startASSEMBLY()        
    linksskroll=Scrollbar(toplist, command=linksListbox.yview)
    linksListbox.pack(side=LEFT, fill=Y)
    linksskroll.pack(side=LEFT, fill=Y)
    linksListbox.config(yscrollcommand=linksskroll.set)
    # <Double-Button-1> – двойной клик левой кнопкой мыши
    linksListbox.bind('<Double-Button-1>', newLink)
    linksListbox.bind('<Return>', newLink)
    root.bind('<Button-1>', exitmenu)
    #toplist.bind('<FocusOut>', exitmenu)
    toplist.after(180000, lambda:toplist.destroy())
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++  
# Смотреть текст строки в колонке
def watch_text(event, columnname):
    pos=columnname.curselection()
    if pos == ():
        mb.showinfo('Не выбран элемент','Нужно выделить элемент левой кнопкой мыши')
        print('Не выбран элемент')
    else:
        columntext=(columnname.get(pos))
        mb.showinfo('', columntext)
    
# Раскрытие меню поиска артикулов
def TMEsearch(event):
    pos=lbox.curselection()    
    if pos == ():
        mb.showinfo('Не выбран элемент','Для открытия меню нужно выделить элемент левой кнопкой мыши')
        print('Не выбран элемент')
    else:
        TMEsearchmenu.post(event.x_root,event.y_root)
        
# Команда на поиск артикула для меню
def TMEsearch_menu_command():
    pos=lbox.curselection()
    searchname=(lbox.get(pos))
    webbrowser.open_new('https://www.tme.eu/ru/katalog/?search='+searchname+'&s_field=1000011&s_order=desc')
  
# Открытием меню
def OPEN_link(event, columnname, menuname):
    pos=columnname.curselection()    
    if pos == ():
        mb.showinfo('Не выбран элемент','Для открытия меню нужно выделить элемент левой кнопкой мыши')
        print('Не выбран элемент')
    else:
        menuname.post(event.x_root,event.y_root)
        
# Срабатывает при нажатии на меню
def OPEN_link_menu_command(columnname):
    pos=columnname.curselection()
    linkname=(columnname.get(pos))
    webbrowser.open_new(linkname) # r'https://'+ 
   
# Раскрытие редактора описания
def editor(event):
    pos=lbox8_link.curselection()    
    if pos == ():
        mb.showinfo('Не выбран элемент','Для открытия меню нужно выделить элемент левой кнопкой мыши')
        print('Не выбран элемент')
    else:
        linkname=(lbox8_link.get(pos))
        if linkname == "":
            editormenu0.post(event.x_root,event.y_root)
        elif linkname[-5:] == '_func':
            mb.showwarning('Возможно только ручное редактирование в БД','Ссылки заканчивающиеся на _func указывают на функцию включения альтернативного сценария сборки описаний, они не редактируются стандартным окном редактора. Для данной функции возможно только ручное редактирование в БД.')
            #editormenufunc.post(event.x_root,event.y_root)
        else:
            editormenu.post(event.x_root,event.y_root)

#
#=====================================++++++++++++++++++++++======редактор шаблонов описаний
def editor_menu_command():
    # выход из топлевела
    def exitmenu():
        editorlevel.destroy()
    def double_scrollbar(*args):
        replacementin.yview(*args)
        replacementout.yview(*args)
    def deleterepl(event):
        indexin=replacementin.curselection()
        indexout=replacementout.curselection()
        if indexin != ():
            masterindex=indexin[0]
        elif indexout != ():
            masterindex=indexout[0]
        else:
            pass
        if masterindex==0:
            pass
        else:
            replacementin.delete(masterindex)
            replacementout.delete(masterindex)
    # ++++++++ активировать меню для ввода+++++++ 
    def inputreplase(event):
        # кнопка >>>
        def confirm():
            confirmin=inputentry.get()
            confirmout=outputentry.get()
            if confirmin == "":
                mb.showwarning("Ошибка", 'В поле "Заменить это" должен быть введен текст')
            else:
                if masterindex == 0:
                    replacementin.insert(END,confirmin)
                    replacementout.insert(END,confirmout)
                else:
                    replacementin.delete(masterindex)
                    replacementout.delete(masterindex)
                    replacementin.insert(masterindex,confirmin)
                    replacementout.insert(masterindex,confirmout)
            print([confirmin,confirmout])
            inputlevel.destroy()
        # LДля нажатия энтер
        def enterconfirm(event):
            confirm()
        # Создаем всплывающее окно   
        inputlevel=Toplevel()
        inputlevel.title('Ввод замен')
        # Убирает рамку вокруг окна
        #inputlevel.overrideredirect(True) 
        x = root.winfo_pointerx()
        y = root.winfo_pointery()
        inputlevel.geometry('+'+str(x)+'+'+str(y)) # '450x140'+
        indexin=replacementin.curselection()
        indexout=replacementout.curselection()
        #print(indexin,indexout)
        # получаем индекс в бокслистах независимо от тыкнутого мышью
        if indexin != ():
            masterindex=indexin[0]
        elif indexout != ():
            masterindex=indexout[0]
        else:
            masterindex=0
        # получаем текст из обоих столбцов
        repltextin=replacementin.get(masterindex)
        repltextout=replacementout.get(masterindex)
        # Создаем элементы
        inputentry=Entry(inputlevel, width=70)
        outputentry=Entry(inputlevel, width=70)
        inputentry.insert(0,repltextin)
        outputentry.insert(0,repltextout)
        enterbutton = Button(inputlevel,text='>>>', command=confirm)
        # Размещение элементов
        Label(inputlevel, text='Заменить это:',font='Arial 8').pack(padx=1, pady=2, anchor=W)
        inputentry.pack(padx=1, pady=1)
        Label(inputlevel, text='На это (поле можно оставить пустым):',font='Arial 8').pack(padx=1, pady=2,anchor=W)
        outputentry.pack(padx=1, pady=2)
        enterbutton.pack(padx=1, pady=2)
        inputentry.bind('<Return>', enterconfirm)
        outputentry.bind('<Return>', enterconfirm)
    def savejson():
        newreplist=[]
        Arepl=replacementin.get(1,END)
        Brepl=replacementout.get(1,END)
        textinget=textin.get(1.0,END)
        textoutget=textout.get(1.0,END)
        # Удаляем перенос строки
        def remove_n(textinout):
            textinout = textinout.rstrip('\n')
            return textinout
        itemsin = remove_n(textinget)
        itemsout = remove_n(textoutget)
        for i in range(len(Arepl)):
            newreplist.append([Arepl[i],Brepl[i]])
        if newreplist == []:
            newjsonitems = [itemsin,itemsout]
            data[1][linkname][0]=newjsonitems[0]
            data[1][linkname][1]=newjsonitems[1]
        else:
            newjsonitems = [itemsin,itemsout,newreplist]   
            #data[1][linkname]=newjsonitems # Эта строка все подтирает 
            data[1][linkname][0]=newjsonitems[0]
            data[1][linkname][1]=newjsonitems[1]
            try:
                data[1][linkname][2]=newjsonitems[2]
            except IndexError:
                data[1][linkname].append(newjsonitems[2])
        
        # Создаем копию БД JSON
        jsonfilename_backup_copy = jsonfilename + '_backup_copy.json'
        if os.path.exists(jsonfilename_backup_copy):
            os.remove(jsonfilename_backup_copy)
            shutil.copy2(jsonfilename, jsonfilename_backup_copy)
        else:
            shutil.copy2(jsonfilename, jsonfilename_backup_copy)
        #os.rename("C://SomeDir/somefile.txt", "C://SomeDir/hello.txt")
        # Сохраняем БД
        with open(jsonfilename,'w') as file:
            json.dump(data, file, indent=2, ensure_ascii=False)
        editorlevel.destroy()
        startASSEMBLY()
    
    # Загружаем базу
    try:
        data = json.load(open(jsonfilename)) # FileNotFoundError:
    except FileNotFoundError:
        print('Файл базы database.json отсутствует, или неверно указанно имя файла')
    #
    # Получаем требуемые переменные
    pos=lbox8_link.curselection()
    linkname=(lbox8_link.get(pos)) # КЛЮЧ !
    # Получаем первую часть шаблона
    jsontextin=(data[1][linkname][0])
    # Получаем вторую часть шаблона
    jsontextout=(data[1][linkname][1])    
    # Основной цвет
    backgroundColor = 'gray80'
    # создаем окно    
    editorlevel=Toplevel()
    editorlevel.title('Редактор шаблона')
    editorlevel.configure(background = backgroundColor)
    #x = root.winfo_pointerx()
    #y = root.winfo_pointery()
    #editorlevel.geometry('180x300'+'+'+str(x)+'+'+str(y))
    #
    # Оставляет рамку вокруг окна
    editorlevel.overrideredirect(False) 
    # Создание элементов
    boxscrollframe=Frame(editorlevel, background = backgroundColor)
    l0 = Label(editorlevel, background = backgroundColor, text='Шаблон по ссылке:')
    linklabel = Label(editorlevel, width=50, background = backgroundColor, text=linkname, font='Arial 14')
    l1 = Label(editorlevel, background = backgroundColor, text='Фраза для начала описания:')
    textin=Text(editorlevel, wrap=WORD, width=75, height=7)
    textin.insert(1.0,jsontextin)
    l2 = Label(editorlevel, background = backgroundColor, text='Фраза для конца описания:')
    textout=Text(editorlevel, wrap=WORD, width=75, height=7)
    textout.insert(1.0,jsontextout)
    l3 = Label(editorlevel, background = backgroundColor, text='Список пар для замен:')
    # Поле с списком замен (заменяемое)
    replacementin=Listbox(editorlevel,  width=50)
    # Скролл
    doublescroll=Scrollbar(boxscrollframe, bg = backgroundColor, command=double_scrollbar)
    # Поле с списком замен (заменитель)
    replacementout=Listbox(boxscrollframe, width=50)
    # Заполнеие замен
    try:
        repllist = data[1][linkname][2]
        for i in repllist:
            replacementin.insert(0,i[0])
            replacementout.insert(0,i[1])
        replacementin.insert(0,"") 
        replacementout.insert(0,"") 
    except IndexError:
        replacementin.insert(0,"") 
        replacementout.insert(0,"")
    # Кнопки
    ok = Button(editorlevel, text='Сохранить', command = savejson)
    cancel = Button(editorlevel, text='Отмена', command=exitmenu)
    #
    l0.grid(row=0, column=0, sticky=SW, padx=4, pady=4 )
    linklabel.grid(row=1, column=0, sticky=NW, columnspan=4, padx=4, pady=4 )
    l1.grid(row=2, column=0, sticky=SW, padx=4, pady=4)
    textin.grid(row=3, column=0, columnspan=4, padx=4, pady=4 )
    l2.grid(row=4, column=0, sticky=SW, padx=4, pady=4 )
    textout.grid(row=5, column=0, columnspan=4, padx=4, pady=4 )
    l3.grid(row=6, column=0, sticky=SW, padx=4, pady=4 )
    replacementin.grid(row=7, column=0,sticky=W+E, columnspan=2, padx=1, pady=4 )
    # Рамка
    boxscrollframe.grid(row=7, column=2, columnspan=2, sticky=W+E, padx=1, pady=4) #, padx=1, pady=4)
    doublescroll.pack(side=LEFT, fill= Y, ipady = 2) #grid(row=0, column=0, sticky=N+S)
    replacementout.pack(side=LEFT, padx=1) #.grid(row=0, column=1, columnspan=2, sticky=W+E)
    ok.grid(row=8, column=2, sticky=E, padx=8, pady=6 )
    cancel.grid(row=8, column=3, sticky=W, padx=8, pady=6 )
    # Для скроллера
    replacementin.config(yscrollcommand=scroll.set)
    # Двойной клик
    replacementin.bind('<Double-Button-1>', inputreplase)
    replacementout.bind('<Double-Button-1>', inputreplase)
    # Удалить кнопкой Del
    replacementin.bind('<Delete>', deleterepl)
    replacementout.bind('<Delete>', deleterepl)
#    
# ++++++==========================++++++++======================Создание нового шаблона
def new_textpat_menu_command():
        # выход из топлевела
    def exitmenu():
        editorlevel.destroy()
    def double_scrollbar(*args):
        replacementin.yview(*args)
        replacementout.yview(*args)
    # Удалить кнопкой DEl
    def deleterepl(event):
        indexin=replacementin.curselection()
        indexout=replacementout.curselection()
        if indexin != ():
            masterindex=indexin[0]
        elif indexout != ():
            masterindex=indexout[0]
        else:
            pass
        if masterindex==0:
            pass
        else:
            replacementin.delete(masterindex)
            replacementout.delete(masterindex)
    # активировать меню для ввода замен
    def inputreplase(event):
        # кнопка >>>
        def confirm():
            confirmin=inputentry.get()
            confirmout=outputentry.get()
            if confirmin == "":
                mb.showwarning("Ошибка", 'В поле "Заменить это" должен быть введен текст')
            else:
                if masterindex == 0:
                    replacementin.insert(END,confirmin)
                    replacementout.insert(END,confirmout)
                else:
                    replacementin.delete(masterindex)
                    replacementout.delete(masterindex)
                    replacementin.insert(masterindex,confirmin)
                    replacementout.insert(masterindex,confirmout)
            print([confirmin,confirmout])
            inputlevel.destroy()
        # LДля нажатия энтер
        def enterconfirm(event):
            confirm()
        # Создаем всплывающее окно   
        inputlevel=Toplevel()
        inputlevel.title('Ввод замен')
        # Убирает рамку вокруг окна
        #inputlevel.overrideredirect(True) 
        x = root.winfo_pointerx()
        y = root.winfo_pointery()
        inputlevel.geometry('+'+str(x)+'+'+str(y)) # '450x140'+
        indexin=replacementin.curselection()
        indexout=replacementout.curselection()
        #print(indexin,indexout)
        # получаем индекс в бокслистах независимо от тыкнутого мышью
        if indexin != ():
            masterindex=indexin[0]
        elif indexout != ():
            masterindex=indexout[0]
        else:
            masterindex=0
        # получаем текст из обоих столбцов
        repltextin=replacementin.get(masterindex)
        repltextout=replacementout.get(masterindex)
        # Создаем элементы
        inputentry=Entry(inputlevel, width=70)
        outputentry=Entry(inputlevel, width=70)
        inputentry.insert(0,repltextin)
        outputentry.insert(0,repltextout)
        enterbutton = Button(inputlevel,text='>>>', command=confirm)
        # Размещение элементов
        Label(inputlevel, text='Заменить это:',font='Arial 8').pack(padx=1, pady=2, anchor=W)
        inputentry.pack(padx=1, pady=1)
        Label(inputlevel, text='На это (поле можно оставить пустым):',font='Arial 8').pack(padx=1, pady=2,anchor=W)
        outputentry.pack(padx=1, pady=2)
        enterbutton.pack(padx=1, pady=2)
        inputentry.bind('<Return>', enterconfirm)
        outputentry.bind('<Return>', enterconfirm)
    # НАжатие кнопки сохранить
    def savejson():
        newTMELink = entryLink.get()
        def saveproces():
            newreplist=[]
            Arepl=replacementin.get(1,END)
            Brepl=replacementout.get(1,END)
            textinget=textin.get(1.0,END)
            textoutget=textout.get(1.0,END)
            # Удаляем перенос строки
            def remove_n(textinout):
                textinout = textinout.rstrip('\n')
                return textinout
            itemsin = remove_n(textinget)
            itemsout = remove_n(textoutget)
            for i in range(len(Arepl)):
                newreplist.append([Arepl[i],Brepl[i]])
            if newreplist == []:
                newjsonitems = [itemsin,itemsout]
            else:
                newjsonitems = [itemsin,itemsout,newreplist]
            data[1][newTMELink]=newjsonitems
            data[0][typeTMEname] = newTMELink
            # Создаем копию БД JSON
            jsonfilename_backup_copy = jsonfilename + '_backup_copy.json'
            if os.path.exists(jsonfilename_backup_copy):
                os.remove(jsonfilename_backup_copy)
                shutil.copy2(jsonfilename, jsonfilename_backup_copy)
            else:
                shutil.copy2(jsonfilename, jsonfilename_backup_copy)
                
            # Сохраняем БД
            with open(jsonfilename,'w') as file:
                json.dump(data, file, indent=2, ensure_ascii=False)
            editorlevel.destroy()
            startASSEMBLY()
        if newTMELink == "":
            mb.showwarning("Ошибка", 'В поле "Ссылка на шаблон" должен быть введен текст')
        else:
            if newTMELink not in [*data[1]]:
                saveproces()
            else:
                answer = mb.askyesno(title="Заменить?", message = 'Ссылка с таким именем уже существует, заменить?')
                if answer == True:
                    saveproces()
    # Переменные                
    # Загружаем базу
    try:
        data = json.load(open(jsonfilename)) # FileNotFoundError:
    except FileNotFoundError:
        print('Файл базы database.json отсутствует, или неверно указанно имя файла')
    # Номер строки в листбоксе
    pos=lbox8_link.curselection()
    #linkname=(lbox8_link.get(pos))
    typeTMEname = lbox7_type.get(pos)
    #
    backgroundColor = 'gray80'
    # создаем окно    
    editorlevel=Toplevel()
    editorlevel.title('Новый шаблон')
    editorlevel.configure(background = backgroundColor)
    x = root.winfo_pointerx()
    y = root.winfo_pointery()
    #editorlevel.geometry('180x300'+'+'+str(x)+'+'+str(y))
    
    # Оставляет рамку вокруг окна
    editorlevel.overrideredirect(False) 
    
    # Создаем элементы
    boxscrollframe=Frame(editorlevel, background = backgroundColor)
    l0 = Label(editorlevel, background = backgroundColor, text='Ссылка на шаблон:')
    entryLink = Entry(editorlevel, width=50)
    #entryLink.insert(0,linkname)
    l1 = Label(editorlevel, background = backgroundColor, text='Фраза для начала описания:')
    textin=Text(editorlevel, width=75, height=7)
    l2 = Label(editorlevel, background = backgroundColor, text='Фраза для конца описания:')
    textout=Text(editorlevel, width=75, height=7)
    l3 = Label(editorlevel, background = backgroundColor, text='Список пар для замен:')
    # Окошко с списком замен (заменяемое)
    replacementin=Listbox(editorlevel,  width=50)
    # Скролл
    doublescroll=Scrollbar(boxscrollframe, bg = backgroundColor, command=double_scrollbar)
    # Окошко с списком замен (заменитель)
    replacementout=Listbox(boxscrollframe, width=50)
    # Заполнение полей замен пустой строкой
    replacementin.insert(0,"") 
    replacementout.insert(0,"")
    # Кнопки
    ok = Button(editorlevel, text='Сохранить', command = savejson)
    cancel = Button(editorlevel, text='Отмена', command=exitmenu)
    # Размещение
    l0.grid(row=0, column=0, sticky=SW, padx=4, pady=4 )
    entryLink.grid(row=1, column=0, sticky=NW, columnspan=4, padx=12, pady=4 )
    l1.grid(row=2, column=0, sticky=SW, padx=4, pady=4)
    textin.grid(row=3, column=0, columnspan=4, padx=4, pady=4 )
    l2.grid(row=4, column=0, sticky=SW, padx=4, pady=4 )
    textout.grid(row=5, column=0, columnspan=4, padx=4, pady=4 )
    l3.grid(row=6, column=0, sticky=SW, padx=4, pady=4 )
    replacementin.grid(row=7, column=0,sticky=W+E, columnspan=2, padx=1, pady=4 )
    # Рамка
    boxscrollframe.grid(row=7, column=2, columnspan=2, sticky=W+E, padx=1, pady=4) #, padx=1, pady=4)
    doublescroll.pack(side=LEFT, fill= Y, ipady = 2) #grid(row=0, column=0, sticky=N+S)
    replacementout.pack(side=LEFT, padx=1) #.grid(row=0, column=1, columnspan=2, sticky=W+E)
    ok.grid(row=8, column=2, sticky=E, padx=8, pady=6 )
    cancel.grid(row=8, column=3, sticky=W, padx=8, pady=6 )
    # кнопки
    ok.grid(row=8, column=1, sticky=E, padx=8, pady=6 )
    cancel.grid(row=8, column=2, sticky=W, padx=8, pady=6 )
    
    # Для скроллера
    replacementin.config(yscrollcommand=scroll.set)
    # Двойной клик
    replacementin.bind('<Double-Button-1>', inputreplase)
    replacementout.bind('<Double-Button-1>', inputreplase)
    # Удалить кнопкой Del
    replacementin.bind('<Delete>', deleterepl)
    replacementout.bind('<Delete>', deleterepl)

# удаление шаблона
def delete_template():
    # Загружаем базу
    try:
        data = json.load(open(jsonfilename)) # FileNotFoundError:
    except FileNotFoundError:
        print('Файл базы database.json отсутствует, или неверно указанно имя файла')
     
    # Номер строки в листбоксе
    pos=lbox8_link.curselection()
    linkname=(lbox8_link.get(pos))
    message = 'Удалить шаблон "'+str(linkname)+'" ? Все ссылки на шаблон заменится на null, будет создана резервная кобия базы данных JSON.'
    delanswer = mb.askyesno(title="Удаление шаблона", message = message)
    if delanswer == True:
        data[1].pop(linkname)
        for i in [*data[0]]:
            if data[0][i]== linkname:
                print(data[0][i])
                data[0][i] = None
        # Создаем копию БД JSON
        jsonfilename_backup_copy = jsonfilename + '_backup_copy.json'
        if os.path.exists(jsonfilename_backup_copy):
            os.remove(jsonfilename_backup_copy)
            shutil.copy2(jsonfilename, jsonfilename_backup_copy)
        else:
            shutil.copy2(jsonfilename, jsonfilename_backup_copy)
            
        # Сохраняем БД
        with open(jsonfilename,'w') as file:
            json.dump(data, file, indent=2, ensure_ascii=False)
        startASSEMBLY()
# О приложении
def about_application():
    inform = " Это программа для получения информации по артикулам товаров с сайта TME (Transfer Multisort Elektronik - https://www.tme.eu/ru/), а также по составлению краткого описания в контексте ТН ВЭД с использованием полученной информации.\n Для получения информации используются команды программного интерфейса(API) сайта TME. Описания составляются на основе базы текстовых шаблонов составляемых пользователем. \n Нажимая ОK ты подтверждаешь передачу своей материальной и интеллектуальной собственности автору программы, а так же даёшь согласие на добровольное чипирование. \n2020 г."
    mb.showinfo('О программе',inform)
        
# Справка
def instructions():
    os.startfile('instructions.pdf')
#
#========================================================================================
# = = = = = = = = = = = =  = = = = = !!! TKINTER !!! = = = = = = = =  = = = = = = == = = = = = =  = = = = = = = >
#=========================================================================================
root =Tk()
root.title("АЛЁ, это Transfer Multisort Elektronik ?") # Заголовок окна
# верхняя менюха
mainmenu = Menu(root) 
helpmenu = Menu(mainmenu,tearoff = 0)
root.config(menu=mainmenu)
mainmenu.add_command(label='Выход', command = quitme)
mainmenu.add_cascade(label='Справка', menu = helpmenu)

helpmenu.add_command(label='О программе',command = about_application)
helpmenu.add_separator()
helpmenu.add_command(label='Инструкция в PDF', command = instructions)


# менюха открытия ссылок  PDF браузере 
TME_PDFlink_menu = Menu(tearoff=0)
#TME_PDFlink_menu.add_command(label='Открыть', command = lambda columnname= lbox4_datasheet : OPEN_link_menu_command(columnname))

# менюха открытия ссылок PICture в браузере 
TME_PIClink_menu = Menu(tearoff=0)
#TME_PIClink_menu.add_command(label='Открыть', command = lambda columnname= lbox5_picture : OPEN_link_menu_command(columnname))


# менюха поиска на ТМЕ
TMEsearchmenu = Menu(tearoff=0)
TMEsearchmenu.add_command(label='Найти на ТМЕ', command=TMEsearch_menu_command)

# менюха ссылок 
linksmenu = Menu(tearoff=0) 
linksmenu.add_command(label='Изменить ссылку на описание', command=links_menu_command)
# менюха ссылок 
linksmenu = Menu(tearoff=0) 
linksmenu.add_command(label='Изменить ссылку на описание', command=links_menu_command)

# меню редактора описаний
editormenu = Menu(tearoff=0)
editormenu.add_command(label='Редактировать шаблон', command=editor_menu_command)
editormenu.add_command(label='Создать новый шаблон', command=new_textpat_menu_command)
editormenu.add_command(label='Удалить', command=delete_template)
# меню только для создания нового описания
editormenu0 = Menu(tearoff=0)
editormenu0.add_command(label='Создать новый шаблон', command=new_textpat_menu_command)

#меню-прудупреждение что невозможно редактировать функцию
#editormenufunc = Menu(tearoff=0)
#editormenufunc.add_cascade(label='Ссылки заканчивающиеся на _func указывают на функцию ')
#editormenufunc.add_cascade(label='включения альтернативного сценария сборки описаний,')
#editormenufunc.add_cascade(label='Они не редактируются стандартным окном редактора. Для данной')
#editormenufunc.add_cascade(label='функции возможно только ручное редактирование в БД.')

# Рамки
visibility_button_frame=Frame()
button_frame=Frame()
f=LabelFrame()
f1=Frame(f)
f1_1=Frame(f)
f2=Frame(f)
f3=Frame(f)
f4=Frame(f)
f5=Frame(f)
f6=Frame(f)
f7=Frame(f)
f8=Frame(f)
f9=Frame(f)
#
# Кнопка отображения колонки артикулов
Art_visibility_button=Button(visibility_button_frame, text="Арт.[x]", bd=0, bg='grey95',font='Arial 7', height=1, command=visibilityArt)
#
# ++++++++++++++++++++++++++++++++++++++++КНОПКА СТАРТА++++++++++++++++++++++++++++++++
start_button=Button(button_frame, text="Запуск", command=startDATA,width=12, font='Arial 10 bold', height=1,bd=4,bg='brown1', activebackground='indianRed3')
# КНОПКА ОТМЕНЫ
#cancel_button=Button(button_frame, text="Отмена", bd=1, command=cancelDATA)
# КНОПКА СБОРКИ ОПИСАНИЯ
#assembly_button=Button(button_frame, text="Составить описание",bd=1,command=startASSEMBLY)
# Ред. базу
editbutton=Button(button_frame, text="Редактировать БД", bd=1,command=DBedit)
# Открыть XLSX
editxlsx=Button(button_frame, text="Открыть XLSX",bd=1,bg='PaleGreen', activebackground='DarkSeaGreen4', command=editXLSX)
#
# Заголовки колонок
label_art =Label(f1,text='Артикулы(A)',height=2)
empty_label=Label(f1_1,height=2)
empty_label_b=Label(button_frame, width=10, height=2)
label_descr =Label(f2,text='Описание(I)',height=2)
label_param =Label(f3,text='Параметры(G)',height=2)
label_weight =Label(f4,text='Вес (кг) (D)',height=2)
label_datasheet =Label(f5,text='Доки(H)',height=2)
label_picture =Label(f6,text='Картинки(E)',height=2)
label_specification =Label(f7,text='Краткое \nописание(C)',height=2)
label_type =Label(f8,text='Тип по ТМЕ(J)',height=2)
label_link =Label(f9,text='Ссылка на\nшаблон-описание',height=2)
#
lbox = Listbox(f1,width=25, height=25,bd=1, selectmode=EXTENDED)
lbox1_descr = Listbox(f2,width=35, height=25,bd=1, selectmode=EXTENDED)# I
lbox2_param = Listbox(f3,width=20, height=25,bd=1, selectmode=EXTENDED)# G
lbox3_weight = Listbox(f4,width=10, height=25,bd=1, selectmode=EXTENDED)# D
lbox4_datasheet = Listbox(f5,width=10, height=25,bd=1, selectmode=EXTENDED)# H
lbox5_picture = Listbox(f6,width=10, height=25,bd=1, selectmode=EXTENDED)# E
lbox6_specification = Listbox(f7,width=20, height=25,bd=1, selectmode=EXTENDED)# C
lbox7_type = Listbox(f8,width=30, height=25,bd=1)# J
lbox8_link = Listbox(f9,width=30, height=25,bd=1)# ССыль
#
scroll = Scrollbar(f1_1,command=on_scrollbar)#lbox.yview)
scrollx = Scrollbar(f1,command=lbox.xview, orient=HORIZONTAL)
scrollx1= Scrollbar(f2,command=lbox1_descr.xview, orient=HORIZONTAL)
scrollx2= Scrollbar(f3,command=lbox2_param.xview, orient=HORIZONTAL)
scrollx3= Scrollbar(f4,command=lbox3_weight, orient=HORIZONTAL)
scrollx4= Scrollbar(f5,command=lbox4_datasheet.xview, orient=HORIZONTAL)
scrollx5= Scrollbar(f6,command=lbox5_picture.xview, orient=HORIZONTAL)
scrollx6= Scrollbar(f7,command=lbox6_specification.xview, orient=HORIZONTAL)
scrollx7= Scrollbar(f8,command=lbox7_type.xview, orient=HORIZONTAL)
scrollx8= Scrollbar(f9,command=lbox8_link.xview, orient=HORIZONTAL)
#
delbutton=Button(button_frame, text="Отчистить", width=11, height=1,bd=1,command=delAll)
# ЛЭЙБЛ ОТОБРАЖАЮЩИЙ ИНФОРМАЦИЮ
infolabel=Label(text=TEXT, width=50, height=3)
# привязка событий
lbox.bind('<Control-v>',ctrlv)
lbox.bind('<Delete>',delList)
lbox.bind('<Control-x>', ctrlx)
lbox.bind('<Return>', startDATAenter)
root.bind('<Return>', startDATAenter)
lbox.bind('<Button-3>', TMEsearch)
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Посмореть текст в информационном окне
lbox1_descr.bind('<Double-Button-1>', lambda event, columnname=lbox1_descr : watch_text(event, columnname))
lbox2_param.bind('<Double-Button-1>', lambda event, columnname=lbox2_param : watch_text(event, columnname))
lbox6_specification.bind('<Double-Button-1>', lambda event, columnname=lbox6_specification : watch_text(event, columnname))
# Открытие ссылок в браузере
TME_PDFlink_menu.add_command(label='Открыть', command = lambda columnname= lbox4_datasheet : OPEN_link_menu_command(columnname))
TME_PIClink_menu.add_command(label='Открыть', command = lambda columnname= lbox5_picture : OPEN_link_menu_command(columnname))

lbox4_datasheet.bind('<Button-3>', lambda event, columnname = lbox4_datasheet, menuname = TME_PDFlink_menu : OPEN_link(event, columnname, menuname))
lbox5_picture.bind('<Button-3>', lambda event, columnname = lbox5_picture, menuname = TME_PIClink_menu  : OPEN_link(event, columnname, menuname))

lbox7_type.bind('<Button-3>', links_menu) 
lbox8_link.bind('<Button-3>', editor) 
# Привязываем событие закрытия главного окна "крестиком" к функции-обработчику close или quit (root.destroy в данном случае):
root.protocol("WM_DELETE_WINDOW", quitme) 
#==============================================================================>
# словарь для кнопок визуализации
visibility_elements={ 2:("Описание[x]","Описание", label_descr, lbox1_descr, scrollx1, f2, clmn_key2,"описание"),3:("Параметры[x]","Параметры", label_param, lbox2_param, scrollx2, f3, clmn_key3,"параметры"),
                    4:("Вес[x]","Вес", label_weight ,lbox3_weight, scrollx3, f4, clmn_key4, "вес"), 5:("Доки[x]","Доки", label_datasheet, lbox4_datasheet, scrollx4, f5, clmn_key5, "доки"),
                    6:("Картинки[x]","Картинки", label_picture, lbox5_picture,scrollx5, f6, clmn_key6, "картинки"), 7:("Кратк.оп-е[x]","Кратк.оп-е", label_specification, lbox6_specification, scrollx6, f7, clmn_key7,"краткопе"),
                    8:("Типы[x]","Типы",label_type,lbox7_type,scrollx7,f8, clmn_key8, "типы"),9:("Ссылка оп-я[x]","Ссылка на описание",label_link,lbox8_link,scrollx8,f9, clmn_key9,"ссылкаопя")}
#===================================================================================================================
#                                      Класс кнопок скрытия-отображения столбцов
# ===================================================================================================================
#visibility_elements={ [i] : ( [0] скрыть, [1] показать, [2] лейбл, [3] бокс, [4] скрол, [5] рамка ) }
class visibility:
    def __init__(self,i):
        self.i = i
        self.button = Button(visibility_button_frame, text=visibility_elements[i][0], bd=0, font='Arial 7', height=1, command=self.visibility_on_off)
        if visibility_elements[self.i][6]=='1':
            self.button['text']=visibility_elements[self.i][0]
            self.button['bg']='grey95'
        else:
            self.button['text']=visibility_elements[self.i][1]
            self.button['bg']='grey90' 
        self.button.pack(side=LEFT, padx=1,pady=2)
        
    def visibility_on_off(self):
        if visibility_elements[self.i][3].winfo_ismapped():
            self.button['text']=visibility_elements[self.i][1]
            self.button['bg']='grey90'
            
            visibility_elements[self.i][2].pack_forget()
            visibility_elements[self.i][3].pack_forget()
            visibility_elements[self.i][4].pack_forget()
            visibility_elements[self.i][5].pack_forget()
        else:
            self.button['text']=visibility_elements[self.i][0]
            self.button['bg']='grey95'
            visibility_elements[self.i][5].pack(side=LEFT)
            visibility_elements[self.i][2].pack(side=TOP)
            visibility_elements[self.i][3].pack(side=TOP, padx=2)
            visibility_elements[self.i][4].pack(fill=X, side=BOTTOM, ipadx=1, padx=2)
    # Установка знчений в словаре с элементами visibility_elements    
    def visibility_config(self): # clmn2.visibility_config()
        if visibility_elements[self.i][3].winfo_ismapped():
            config['COLUMNS'][visibility_elements[self.i][7]] = '1'
        else:
            config['COLUMNS'][visibility_elements[self.i][7]] = '0'

#
#=========================        ================================================================        ================
# +++++++++++++++++++++= = == = = = = = = = = Размещение ВИДЖЕТОВ = = = = = = = = = = = = = = == = = = = = = = = = = = = = = >
#=========================         ===============================================================        ================
visibility_button_frame.pack()    
Art_visibility_button.pack(side=LEFT,padx=1)
# Кнопки видимости колонок от класса
clmn2=visibility(2)
clmn3=visibility(3)
clmn4=visibility(4)
clmn5=visibility(5)
clmn6=visibility(6)
clmn7=visibility(7)
clmn8=visibility(8)
clmn9=visibility(9)
button_frame.pack()
# КНОПКА СТАРТА размещение
start_button.pack(side=LEFT,pady=20,padx=10)
# Кнопка отчистки (раньше КНОПКА ОТМЕНЫ)
delbutton.pack(side=LEFT,pady=3,padx=3)
# КНОПКА СБОРКИ ОПИСАНИЯ размещение
#assembly_button.pack(side=LEFT,pady=10,padx=55)
# Пространство
empty_label_b.pack(side=LEFT)
# Ред. базу, размещение
editbutton.pack(side=LEFT,anchor=E,padx=3)
# Открыть XLSC
editxlsx.pack(side=LEFT, anchor=E, padx=3)
# Рамка в которой размещены вертикальные рамки
f.pack(expand=1, padx=15,pady=10)
#f.pack_propagate(0) # - исщезновение виджета

# КОЛОНКА 1
if clmn_key1 == '1':
    f1.pack(side=LEFT) # Вертикальная рамка
    label_art.pack()
    lbox.pack(side=TOP)
    scrollx.pack(fill=X, side=BOTTOM, ipadx=50, padx=2)
    f1_1.pack(side=LEFT,ipady=8)# Рамка для прокрутки
    empty_label.pack()
    scroll.pack(side=TOP, anchor=N, ipady=178)

# КОЛОНКА 2
if clmn_key2 == '1':
    f2.pack(side=LEFT)
    label_descr.pack()
    lbox1_descr.pack(side=TOP, padx=2)
    scrollx1.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 3
if clmn_key3 == '1':
    f3.pack(side=LEFT)
    label_param.pack()
    lbox2_param.pack(side=TOP, padx=2)
    scrollx2.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 4
if clmn_key4 == '1':   
    f4.pack(side=LEFT)
    label_weight.pack()
    lbox3_weight.pack(side=TOP, padx=2)
    scrollx3.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 5
if clmn_key5 == '1':    
    f5.pack(side=LEFT)
    label_datasheet.pack()
    lbox4_datasheet.pack(side=TOP, padx=2)
    scrollx4.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 6
if clmn_key6 == '1':   
    f6.pack(side=LEFT)
    label_picture.pack()
    lbox5_picture.pack(side=TOP, padx=2)
    scrollx5.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 7
if clmn_key7 == '1':    
    f7.pack(side=LEFT)
    label_specification.pack()
    lbox6_specification.pack(side=TOP, padx=2)
    scrollx6.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 8
if clmn_key8 == '1':      
    f8.pack(side=LEFT)
    label_type.pack()
    lbox7_type.pack(side=TOP, padx=2)
    scrollx7.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# КОЛОНКА 9
if clmn_key9 == '1':     
    f9.pack(side=LEFT)
    label_link.pack()
    lbox8_link.pack(side=TOP, padx=2)
    scrollx8.pack(fill=X, side=BOTTOM, ipadx=1, padx=2)

# Проверка условия для кнопки видимости Артикулы
if clmn_key1== '1':
    Art_visibility_button['text']= "Арт.[x]"
    Art_visibility_button['bg']= 'grey95'       
else:
    Art_visibility_button['text']= "Артикулы"
    Art_visibility_button['bg']='grey90'

# Вне рамки 
#delbutton.pack(expand=1,pady=5)
infolabel.pack()
# Для скроллеров
lbox.config(yscrollcommand=scroll.set)
lbox.config(xscrollcommand=scrollx.set)
# заполняем поля
get_xlsxdata()
root.update_idletasks() # обновляем данные об окне
root.geometry(app_geometry)
#print(root.geometry()) # получаем параметры окна (стандарт 1504x661+25+25)
#print(app_geometry)
root.mainloop()